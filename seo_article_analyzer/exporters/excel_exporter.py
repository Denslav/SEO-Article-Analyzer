from __future__ import annotations

from pathlib import Path
import math

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class ExcelExporter:
    def export(self, report: dict[str, list[list]], xlsx_path: str) -> None:
        path = Path(xlsx_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        for sheet_name, rows in report.items():
            worksheet = workbook.create_sheet(title=self._safe_sheet_name(sheet_name))
            for row in rows:
                worksheet.append(row)
            self._style(worksheet, original_sheet_name=sheet_name)

        workbook.save(path)

    def _style(self, worksheet, original_sheet_name: str = "") -> None:
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True)
        wrap = Alignment(wrap_text=True, vertical="top")

        if worksheet.max_row >= 1:
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = wrap

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = wrap

        if original_sheet_name == "Final SEO Brief":
            self._style_final_seo_brief(worksheet)
        else:
            self._style_default_sheet(worksheet)

        worksheet.freeze_panes = "A2"

    def _style_default_sheet(self, worksheet) -> None:
        for column_index, column_cells in enumerate(worksheet.columns, start=1):
            max_length = 10
            for cell in column_cells:
                value = str(cell.value or "")
                max_length = max(max_length, min(len(value), 75))
            worksheet.column_dimensions[get_column_letter(column_index)].width = max_length + 2

        self._autofit_row_heights(worksheet, max_height=160)

    def _style_final_seo_brief(self, worksheet) -> None:
        # Финальный бриф почти всегда состоит из двух колонок:
        # A - название блока, B - длинный текст. Поэтому задаём ширины вручную.
        worksheet.column_dimensions["A"].width = 30
        worksheet.column_dimensions["B"].width = 95

        # Заголовок
        worksheet.row_dimensions[1].height = 24

        for row_index in range(2, worksheet.max_row + 1):
            label = str(worksheet.cell(row=row_index, column=1).value or "")
            value = str(worksheet.cell(row=row_index, column=2).value or "")

            # Подсчёт высоты по реальному количеству строк + примерному переносу по ширине колонки.
            explicit_lines = value.count("\n") + 1 if value else 1
            wrapped_lines = 0

            for part in value.split("\n") if value else [""]:
                # Колонка B шириной 95 символов. Делаем запас, чтобы текст не обрезался.
                wrapped_lines += max(1, math.ceil(len(part) / 88))

            lines = max(explicit_lines, wrapped_lines)

            # Для больших блоков, таких как FAQ или Структура статьи, даём больше места.
            if label in {"Дополнительные ключи", "Сущности / термины", "LSI / тематические слова", "FAQ-вопросы", "Структура статьи"}:
                height = min(max(24, lines * 16), 420)
            else:
                height = min(max(22, lines * 15), 130)

            worksheet.row_dimensions[row_index].height = height

        # Чуть улучшим читаемость финального брифа.
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=2):
            row[0].font = Font(bold=True)
            row[0].alignment = Alignment(wrap_text=True, vertical="top")
            row[1].alignment = Alignment(wrap_text=True, vertical="top")

    def _autofit_row_heights(self, worksheet, max_height: int = 160) -> None:
        # Excel не всегда сам раскрывает строки при wrap_text,
        # поэтому задаём высоту примерно по количеству переносов.
        column_widths = {
            index: worksheet.column_dimensions[get_column_letter(index)].width or 12
            for index in range(1, worksheet.max_column + 1)
        }

        for row_index in range(1, worksheet.max_row + 1):
            max_lines = 1

            for column_index in range(1, worksheet.max_column + 1):
                value = str(worksheet.cell(row=row_index, column=column_index).value or "")
                if not value:
                    continue

                width = max(8, int(column_widths.get(column_index, 12)))
                lines = 0
                for part in value.split("\n"):
                    lines += max(1, math.ceil(len(part) / max(8, width - 2)))

                max_lines = max(max_lines, lines)

            worksheet.row_dimensions[row_index].height = min(max(18, max_lines * 15), max_height)

    def _safe_sheet_name(self, name: str) -> str:
        for char in ["\\", "/", "*", "[", "]", ":", "?"]:
            name = name.replace(char, " ")
        return name[:31]
